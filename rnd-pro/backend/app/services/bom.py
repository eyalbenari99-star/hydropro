"""BOM XLSX + Panel-schedule PDF generation."""
import io
from typing import Dict, Any


def generate_bom_xlsx(spec: Dict[str, Any]) -> bytes:
    """Build a styled BOM XLSX from a CabinetSpecV2 dict."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'BOM'
    headers = ['Ref', 'Description', 'Manufacturer', 'Part No', 'Rating', 'Poles', 'Qty', 'Unit Price', 'Total']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill('solid', fgColor='1F4E79')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    row = 2
    for sec in spec.get('sections', []) or []:
        sec_row = row
        ws.cell(row=row, column=1, value=sec.get('label', ''))
        ws.cell(row=sec_row, column=1).font = Font(bold=True, italic=True, size=11)
        ws.cell(row=sec_row, column=1).fill = PatternFill('solid', fgColor='DDEBF7')
        row += 1
        for rail in sec.get('rails', []) or []:
            for item in rail.get('items', []) or []:
                ws.cell(row=row, column=1, value=item.get('ref', ''))
                ws.cell(row=row, column=2, value=item.get('label', ''))
                ws.cell(row=row, column=3, value=item.get('manufacturer', ''))
                ws.cell(row=row, column=4, value=item.get('partNo', ''))
                ws.cell(row=row, column=5, value=item.get('rating', ''))
                ws.cell(row=row, column=6, value=item.get('poles', ''))
                ws.cell(row=row, column=7, value=1)
                ws.cell(row=row, column=8, value=0)
                ws.cell(row=row, column=9, value=f'=G{row}*H{row}')
                row += 1

    # Totals
    if row > 2:
        ws.cell(row=row+1, column=8, value='TOTAL').font = Font(bold=True)
        ws.cell(row=row+1, column=9, value=f'=SUM(I2:I{row-1})').font = Font(bold=True)

    widths = [10, 38, 16, 26, 12, 8, 8, 12, 12]
    for c, w_ in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w_
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_panel_pdf(spec: Dict[str, Any]) -> bytes:
    """Build a panel schedule PDF via reportlab. Falls back to minimal PDF if reportlab missing."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.units import mm
    except ImportError:
        return _minimal_pdf(spec.get('title', 'Cabinet'))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            topMargin=15*mm, bottomMargin=15*mm,
                            leftMargin=12*mm, rightMargin=12*mm)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph(spec.get('title', 'Cabinet'), styles['Title']))
    if spec.get('subtitle'):
        story.append(Paragraph(spec['subtitle'], styles['Heading3']))
    story.append(Spacer(1, 8*mm))

    data = [['Ref', 'Description', 'Manufacturer', 'Part No', 'Rating', 'Poles']]
    for sec in spec.get('sections', []) or []:
        for rail in sec.get('rails', []) or []:
            for item in rail.get('items', []) or []:
                data.append([
                    str(item.get('ref', '')),
                    str(item.get('label', '')),
                    str(item.get('manufacturer', '')),
                    str(item.get('partNo', '')),
                    str(item.get('rating', '')),
                    str(item.get('poles', '')),
                ])
    if len(data) == 1:
        data.append(['-', 'No devices in spec', '', '', '', ''])

    tbl = Table(data, repeatRows=1, colWidths=[20*mm, 80*mm, 30*mm, 50*mm, 25*mm, 18*mm])
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('GRID',       (0,0), (-1,-1), 0.25, colors.HexColor('#9ca3af')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f3f4f6')]),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(tbl)
    doc.build(story)
    return buf.getvalue()


def _as_payload(spec: Any) -> Dict[str, Any]:
    if hasattr(spec, 'model_dump'):
        return spec.model_dump(by_alias=True, exclude_none=True)
    if isinstance(spec, dict):
        return spec
    raise TypeError('Cabinet specification must be a mapping or Pydantic model')


def build_bom_xlsx(spec: Any) -> bytes:
    """Compatibility entrypoint used by the typed API route."""
    return generate_bom_xlsx(_as_payload(spec))


def build_wire_pdf(spec: Any) -> bytes:
    """Compatibility entrypoint used by the typed API route."""
    return generate_panel_pdf(_as_payload(spec))


def _minimal_pdf(title: str) -> bytes:
    """Tiny valid PDF when reportlab is unavailable."""
    safe = title.replace('(', '').replace(')', '')[:60]
    body = f"BT /F1 24 Tf 50 700 Td ({safe}) Tj ET"
    body_b = body.encode('latin-1', errors='replace')
    lines = [
        b'%PDF-1.4',
        b'1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj',
        b'2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj',
        b'3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 595 842]/Contents 4 0 R/Resources<</Font<</F1<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>>>>>>>endobj',
        b'4 0 obj<</Length ' + str(len(body_b)).encode() + b'>>stream',
        body_b,
        b'endstream endobj',
        b'xref',
        b'0 5',
        b'0000000000 65535 f',
        b'0000000010 00000 n',
        b'0000000060 00000 n',
        b'0000000110 00000 n',
        b'0000000220 00000 n',
        b'trailer<</Size 5/Root 1 0 R>>',
        b'startxref 320',
        b'%%EOF',
    ]
    return b'\n'.join(lines)
